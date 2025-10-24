#!/usr/bin/env python3
"""
Streamlit app to generate CSV and Excel files containing years from birth year to age 100,
with each year split into four seasons, and each season split into months.
Starts from the user's birth month and continues chronologically.
Includes Japanese reign period (era), Chinese zodiac, and Western zodiac for each month.
Features formatted Excel output with colored cells for seasons, and CSV output.
"""

import streamlit as st
import csv
import re
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


def get_japanese_era(year, month):
    """
    Returns the Japanese era name and reign year based on the year and month.
    Format: "Era Name Year" (e.g., "Showa 38", "Heisei 1")
    Modern eras: Meiji, Taisho, Showa, Heisei, Reiwa
    """
    month_num = {
        'January': 1, 'February': 2, 'March': 3, 'April': 4,
        'May': 5, 'June': 6, 'July': 7, 'August': 8,
        'September': 9, 'October': 10, 'November': 11, 'December': 12
    }
    
    m = month_num[month]
    
    # Modern Japanese eras with exact dates and reign years
    if year < 1868:
        return "Pre-Meiji"
    elif year == 1868:
        if m < 9:
            return "Pre-Meiji"
        else:
            return "Meiji 1"
    elif year <= 1911:
        era_year = year - 1867
        return f"Meiji {era_year}"
    elif year == 1912:
        if m <= 7:
            era_year = year - 1867
            return f"Meiji {era_year}"
        else:
            return "Taisho 1"
    elif year <= 1925:
        era_year = year - 1911
        return f"Taisho {era_year}"
    elif year == 1926:
        if m <= 12:
            era_year = year - 1911
            return f"Taisho {era_year}"
        else:
            return "Showa 1"
    elif year <= 1988:
        era_year = year - 1925
        return f"Showa {era_year}"
    elif year == 1989:
        if m <= 1:
            era_year = year - 1925
            return f"Showa {era_year}"
        else:
            return "Heisei 1"
    elif year <= 2018:
        era_year = year - 1988
        return f"Heisei {era_year}"
    elif year == 2019:
        if m <= 4:
            era_year = year - 1988
            return f"Heisei {era_year}"
        else:
            return "Reiwa 1"
    else:
        era_year = year - 2018
        return f"Reiwa {era_year}"


def get_chinese_zodiac(year, month):
    """
    Returns the Chinese zodiac animal based on the year and month.
    """
    zodiac_animals = ['Rat', 'Ox', 'Tiger', 'Rabbit', 'Dragon', 'Snake',
                      'Horse', 'Goat', 'Monkey', 'Rooster', 'Dog', 'Wild Boar']
    
    zodiac_index = (year - 1900) % 12
    zodiac_animal = zodiac_animals[zodiac_index]
    zodiac_year = zodiac_index + 1
    
    return f"{zodiac_animal} Year {zodiac_year}"


def get_western_zodiac(month):
    """
    Returns the Western zodiac sign based on the month.
    """
    western_zodiac = {
        'January': 'Capricorn/Aquarius',
        'February': 'Aquarius/Pisces',
        'March': 'Pisces/Aries',
        'April': 'Aries/Taurus',
        'May': 'Taurus/Gemini',
        'June': 'Gemini/Cancer',
        'July': 'Cancer/Leo',
        'August': 'Leo/Virgo',
        'September': 'Virgo/Libra',
        'October': 'Libra/Scorpio',
        'November': 'Scorpio/Sagittarius',
        'December': 'Sagittarius/Capricorn'
    }
    
    return western_zodiac.get(month, 'Unknown')


def get_month_fill(month):
    """
    Returns the appropriate fill color for a given month.
    Spring months (March, April, May): Green
    Autumn months (September, October, November): Light Red
    Summer and Winter months: No fill
    """
    spring_months = ['March', 'April', 'May']
    autumn_months = ['September', 'October', 'November']
    
    if month in spring_months:
        # Green fill - theme 9 with tint
        return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif month in autumn_months:
        # Light red fill - theme 5 with tint
        return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    else:
        # No fill for summer and winter
        return None


def generate_data(name, birth_year, birth_month):
    """
    Generates the data rows and returns them along with sanitized name.
    This is used by both CSV and Excel generation functions.
    """
    sanitized_name = re.sub(r'[^a-zA-Z0-9_-]', '_', name).lower()
    end_year = birth_year + 100
    
    season_map = {
        'January': 'winter',
        'February': 'winter',
        'March': 'spring',
        'April': 'spring',
        'May': 'spring',
        'June': 'summer',
        'July': 'summer',
        'August': 'summer',
        'September': 'autumn',
        'October': 'autumn',
        'November': 'autumn',
        'December': 'winter',
    }
    
    months = ['January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December']
    
    birth_month_index = months.index(birth_month)
    rotated_months = months[birth_month_index:] + months[:birth_month_index]
    
    # Store all rows
    all_rows = []
    
    # Write header
    header = ['Year', 'Age', 'Season', 'Month', 'Japanese Era', 'Chinese Zodiac', 'Western Zodiac']
    all_rows.append(header)
    
    current_year = birth_year
    month_count = 0
    total_months = (end_year - birth_year + 1) * 12
    
    while month_count < total_months:
        years_passed = month_count // 12
        age = years_passed
        
        month_index = month_count % 12
        month = rotated_months[month_index]
        season = season_map[month]
        
        japanese_era = get_japanese_era(current_year, month)
        chinese_zodiac = get_chinese_zodiac(current_year, month)
        western_zodiac = get_western_zodiac(month)
        
        row = [current_year, age, season, month, japanese_era, chinese_zodiac, western_zodiac]
        all_rows.append(row)
        
        if month == 'December':
            current_year += 1
        
        month_count += 1
    
    return sanitized_name, all_rows


def generate_csv(sanitized_name, all_rows):
    """
    Generates CSV data in memory and returns it as a string.
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    for row in all_rows:
        writer.writerow(row)
    
    csv_data = output.getvalue()
    output.close()
    
    return csv_data


def generate_excel(sanitized_name, all_rows):
    """
    Generates Excel file in memory and returns it as bytes.
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{sanitized_name}_chronology"
    
    # Write header
    header = all_rows[0]
    ws.append(header)
    
    # Style header row
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    
    # Write data rows
    for row_idx, row in enumerate(all_rows[1:], start=2):
        ws.append(row)
        
        # Apply formatting to the row based on month
        month = row[3]  # Month is in column D (index 3)
        fill = get_month_fill(month)
        if fill:
            # Apply fill to entire row
            for col in range(1, 8):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def main():
    st.set_page_config(page_title="Life Chronology Generator", page_icon="📅", layout="centered")
    
    st.title("📅 Life Chronology Generator")
    st.markdown("Generate your personalized chronology from birth to age 100")
    
    # User inputs
    name = st.text_input("**Your name:**", placeholder="e.g., Michael")
    
    col1, col2 = st.columns(2)
    
    with col1:
        birth_year = st.number_input("**Birth year:**", min_value=1868, max_value=2100, value=1990, step=1)
    
    with col2:
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
        birth_month = st.selectbox("**Birth month:**", months)
    
    # Generate button
    if st.button("🚀 Generate Chronology", type="primary", use_container_width=True):
        if not name:
            st.error("⚠️ Please enter your name")
        else:
            with st.spinner("Generating..."):
                sanitized_name, all_rows = generate_data(name, birth_year, birth_month)
                
                st.success(f"✅ **Files Generated Successfully for {name}!**")
                
                # Show stats
                end_year = birth_year + 100
                total_rows = len(all_rows) - 1
                st.info(f"📊 Covering {end_year - birth_year + 1} years ({total_rows} rows)")
                
                st.markdown("---")
                
                # Download section with both formats
                st.markdown("### 📥 Download Your Files")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    csv_data = generate_csv(sanitized_name, all_rows)
                    csv_filename = f"{sanitized_name}_chronology.csv"
                    st.download_button(
                        label="⬇️ DOWNLOAD CSV",
                        data=csv_data,
                        file_name=csv_filename,
                        mime="text/csv",
                        use_container_width=True,
                        type="primary"
                    )
                
                with col2:
                    excel_buffer = generate_excel(sanitized_name, all_rows)
                    xlsx_filename = f"{sanitized_name}_chronology.xlsx"
                    st.download_button(
                        label="⬇️ DOWNLOAD XLSX",
                        data=excel_buffer.getvalue(),
                        file_name=xlsx_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )
                
                st.markdown("")
                
                # Info about formatting
                st.info("💡 **Formatting:** Spring months (Mar-May) are shaded green, Autumn months (Sep-Nov) are shaded light red (XLSX only)")
                
                st.markdown("")
                
                # Compact preview
                st.markdown("### 📋 Preview")
                
                # Show first 5 rows
                st.markdown("**First 5 rows:**")
                preview_first = all_rows[:6]
                for row in preview_first:
                    st.text(', '.join(map(str, row)))
                
                st.markdown("**...**")
                
                # Show last 3 rows up to present day
                current_year = datetime.now().year
                current_month = datetime.now().month
                month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                              'July', 'August', 'September', 'October', 'November', 'December']
                current_month_name = month_names[current_month - 1]
                
                st.markdown(f"**Last 3 rows (up to {current_month_name} {current_year}):**")
                
                # Find rows up to present day
                present_day_rows = []
                for row in all_rows[1:]:
                    if row[0] <= current_year:
                        present_day_rows.append(row)
                
                # Show last 3 rows
                if len(present_day_rows) > 3:
                    preview_last = present_day_rows[-3:]
                else:
                    preview_last = present_day_rows
                
                for row in preview_last:
                    st.text(', '.join(map(str, row)))
    
    # Compact info at bottom
    with st.expander("ℹ️ About"):
        st.markdown("""
        Generates CSV and Excel files with: Years, Age, Seasons, Months, Japanese Era, Chinese Zodiac, Western Zodiac
        
        Ages increment on your birthday anniversary. Covers 100 years from your birth month.
        
        **Excel Formatting:**
        - Spring months (March, April, May): Green background
        - Autumn months (September, October, November): Light red background
        - Summer and Winter months: No background color
        
        **CSV Format:**
        - Plain text format, no formatting
        - Compatible with all spreadsheet applications
        """)


if __name__ == "__main__":
    main()

